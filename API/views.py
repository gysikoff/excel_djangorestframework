from rest_framework.response import Response
from rest_framework.viewsets import ViewSet
from .serializers import UploadExcelSerializer
from rest_framework import status

from openpyxl import load_workbook


class UploadExcelViewSet(ViewSet):
    serializer_class = UploadExcelSerializer

    def list(self, request):
        return Response("upload excel file")

    def create(self, request):
        # Gets request content
        excel_file = request.FILES.get('excel_file')
        colums_raw = request.POST.get('columns')
        content_type = excel_file.content_type

        # Handles wrong file extension
        if content_type != "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet":
            return Response({"detail": "Wrong file extension"}, status=status.HTTP_415_UNSUPPORTED_MEDIA_TYPE)
        
        columns = list(colums_raw.split(', '))
        response = {'file': excel_file.name, 'summary': []} # Creates response object

        cell_value_list = []
        
        wb = load_workbook(excel_file) # Loads excel file
        append_list = False
        # Reads values in every column of excel file
        for ws in wb.worksheets:
            for col in ws.iter_cols():
                append_list = False
                for cell in col:
                    if cell.value is not None:
                        if not append_list:
                            if str(cell.value).strip() in columns: # Finds a cell via cell.value provided in request param column
                                col_name = str(cell.value).strip()
                                append_list = True # Sets the boolean to true so every float or int from cells below is added to a list
                        else:
                            if type(cell.value) == int or type(cell.value) == float:
                                cell_value_list.append(cell.value)
                if append_list:
                    search = {'column': col_name, 'sum': round(sum(cell_value_list), 2), 'avg': round(avg(cell_value_list), 2)} # Creates result object to be outputed in response
                    response['summary'].append(search)
                    cell_value_list = []

        return Response(response)

def sum(list):
    sum = 0
    for x in list:
        sum += x

    return sum

def avg(list):
    sum = 0
    i = 0
    for x in list:
        sum +=x
        i += 1

    return sum/i